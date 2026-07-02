#!/usr/bin/env python3
"""
Transmission Hub — Import Dữ Liệu Tự Động
Thay thế cho `npm run import:samples`.

Flow:
  1. Tìm file Excel mới nhất theo từng prefix trong WATCH_DIR
  2. Xóa toàn bộ dữ liệu cũ trên Supabase (CASCADE tự động xóa tất cả bảng con)
  3. Upload dữ liệu mới lên Supabase và chạy audit + reclaim rule engine
  4. Chuyển tất cả file xlsx (mới + cũ hơn) vào thư mục ARCHIVE_SUBDIR

Cài đặt thư viện:
    pip install openpyxl supabase python-dotenv

Chạy thủ công:
    python import_data.py

Cài Task Scheduler (Windows):
    Program: python
    Arguments: "D:\\path\\to\\import_data.py"
    Start in: D:\\path\\to\\
"""

# ==============================================================================
# CONFIG — Chỉnh các giá trị này trước khi triển khai
# ==============================================================================

# Đường dẫn thư mục chứa file xlsx.
# Để rỗng ("") để dùng thư mục chứa file script này.
WATCH_DIR: str = ""

# Tên thư mục lưu trữ file đã import (tạo tự động nếu chưa có).
# Đường dẫn tương đối so với WATCH_DIR.
ARCHIVE_SUBDIR: str = "imported"

# Supabase URL và Service Role Key.
# Điền trực tiếp ở đây, HOẶC để rỗng và đặt vào file .env trong WATCH_DIR:
#   VITE_SUPABASE_URL=https://xxx.supabase.co
#   SUPABASE_SERVICE_ROLE_KEY=eyJ...
SUPABASE_URL: str = ""
SUPABASE_SERVICE_ROLE_KEY: str = ""

# Số dòng tối đa mỗi lần ghi lên Supabase (giảm nếu gặp lỗi timeout)
CHUNK_SIZE: int = 500

# Nhãn batch (hiển thị trong bảng import_batches)
BATCH_SOURCE_LABEL: str = "auto-import"

# Ghi log ra file (để rỗng nếu chỉ cần in ra console).
# Ví dụ: r"C:\Logs\transmission_hub.log"
LOG_FILE: str = ""

# ==============================================================================
# Phần còn lại không cần chỉnh
# ==============================================================================

import glob
import ipaddress
import logging
import os
import shutil
import sys
from datetime import datetime, timezone
from pathlib import Path

# ---------------------------------------------------------------------------
# Logging
# ---------------------------------------------------------------------------

def _setup_logging() -> None:
    fmt = "%(asctime)s  %(levelname)-8s  %(message)s"
    handlers: list[logging.Handler] = [logging.StreamHandler(sys.stdout)]
    if LOG_FILE:
        handlers.append(logging.FileHandler(LOG_FILE, encoding="utf-8"))
    logging.basicConfig(level=logging.INFO, format=fmt, handlers=handlers)

_setup_logging()
log = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Resolve WATCH_DIR
# ---------------------------------------------------------------------------

_SCRIPT_DIR = str(Path(__file__).resolve().parent)
_watch_dir = Path(WATCH_DIR) if WATCH_DIR else Path(_SCRIPT_DIR)

# ---------------------------------------------------------------------------
# Load credentials từ .env nếu chưa điền ở CONFIG
# ---------------------------------------------------------------------------

_supabase_url = SUPABASE_URL
_service_role_key = SUPABASE_SERVICE_ROLE_KEY

if not _supabase_url or not _service_role_key:
    try:
        from dotenv import load_dotenv
        load_dotenv(_watch_dir / ".env")
        load_dotenv()
    except ImportError:
        pass

    if not _supabase_url:
        _supabase_url = (
            os.environ.get("VITE_SUPABASE_URL")
            or os.environ.get("SUPABASE_URL", "")
        )
    if not _service_role_key:
        _service_role_key = os.environ.get("SUPABASE_SERVICE_ROLE_KEY", "")

if not _supabase_url or not _service_role_key:
    log.error(
        "Chưa cấu hình Supabase. Điền SUPABASE_URL và SUPABASE_SERVICE_ROLE_KEY "
        "vào phần CONFIG đầu file, hoặc tạo file .env trong thư mục:\n  %s",
        _watch_dir,
    )
    sys.exit(1)

# ---------------------------------------------------------------------------
# Kiểm tra thư viện
# ---------------------------------------------------------------------------

try:
    import openpyxl  # noqa: F401
except ImportError:
    log.error("Thiếu thư viện openpyxl. Cài: pip install openpyxl")
    sys.exit(1)

try:
    from supabase import create_client
except ImportError:
    log.error("Thiếu thư viện supabase. Cài: pip install supabase")
    sys.exit(1)

_supabase = create_client(_supabase_url, _service_role_key)

# ==============================================================================
# Chuẩn hóa giá trị từ Excel → Postgres
# ==============================================================================

def _to_str(value) -> "str | None":
    if value is None:
        return None
    s = str(value).strip()
    return s or None


def _to_int(value) -> "int | None":
    if value is None:
        return None
    s = str(value).strip() if isinstance(value, str) else value
    if isinstance(s, str) and not s:
        return None
    try:
        return int(float(str(s)))
    except (ValueError, TypeError):
        return None


def _to_ts(value) -> "str | None":
    """Excel timestamp → 'YYYY-MM-DD HH:MM:SS' string (Postgres parses as timestamptz)."""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    s = str(value).strip()
    return s or None


# ==============================================================================
# Đọc file Excel
# ==============================================================================

def _find_latest(prefix: str) -> str:
    """Trả về đường dẫn đầy đủ của file xlsx mới nhất khớp với prefix."""
    files = sorted(glob.glob(str(_watch_dir / f"{prefix}*.xlsx")))
    if not files:
        raise FileNotFoundError(
            f'Không tìm thấy file "{prefix}*.xlsx" trong {_watch_dir}'
        )
    return files[-1]


def _find_all(prefix: str) -> list:
    return glob.glob(str(_watch_dir / f"{prefix}*.xlsx"))


def _read_sheet(filepath: str, sheet_name: str) -> list:
    """Đọc một sheet Excel, trả về list of dicts (key = header)."""
    import openpyxl as xl
    wb = xl.load_workbook(filepath, data_only=True, read_only=True)
    if sheet_name not in wb.sheetnames:
        wb.close()
        raise ValueError(
            f'Sheet "{sheet_name}" không tồn tại trong {os.path.basename(filepath)}'
        )
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return []
    headers = [
        str(h).strip() if h is not None else f"_col{i}"
        for i, h in enumerate(rows[0])
    ]
    result = []
    for row in rows[1:]:
        if all(v is None for v in row):
            continue  # bỏ qua dòng hoàn toàn trống
        result.append({
            headers[i]: (row[i] if i < len(row) else None)
            for i in range(len(headers))
        })
    return result


# ==============================================================================
# Supabase helpers
# ==============================================================================

def _insert_chunked(table: str, rows: list) -> None:
    if not rows:
        return
    for i in range(0, len(rows), CHUNK_SIZE):
        chunk = rows[i : i + CHUNK_SIZE]
        try:
            _supabase.table(table).insert(chunk).execute()
        except Exception as exc:
            raise RuntimeError(
                f"Insert vào {table} thất bại (dòng {i}–{i + len(chunk) - 1}): {exc}"
            ) from exc
        log.info("  %s  [%d/%d]", table, min(i + CHUNK_SIZE, len(rows)), len(rows))


# ==============================================================================
# Biến đổi dữ liệu: Excel → DB payload
# ==============================================================================

def _inv_row(row: dict, bid: str) -> dict:
    return {
        "import_batch_id": bid,
        "device_name":     _to_str(row.get("Device_Name")),
        "device_ip":       _to_str(row.get("Device_IP")),
        "vendor":          _to_str(row.get("Vendor")),
        "loopback_ip":     _to_str(row.get("Loopback_IP")),
        "interface_name":  _to_str(row.get("Interface")),
        "vlan_id":         _to_int(row.get("VLAN_ID")),
        "vlan_description":_to_str(row.get("VLAN_Description")),
        "vrf_instance":    _to_str(row.get("VRF_Instance")),
        "ip_address":      _to_str(row.get("IP_Address")),
        "prefix_length":   _to_int(row.get("Prefix_Length")),
        "gateway":         _to_str(row.get("Gateway")),
        "physical_port":   _to_str(row.get("Physical_Port")),
        "port_description":_to_str(row.get("Port_Description")),
        "service_type":    _to_str(row.get("Service_Type")),
        "admin_state":     _to_str(row.get("Intf_Adm_State")),
        "oper_state":      _to_str(row.get("Intf_Opr_State")),
        "static_routes":   _to_str(row.get("Static_Routes")),
        "notes":           _to_str(row.get("Notes")),
        "status":          _to_str(row.get("Status")),
    }


def _ospf_intf_row(row: dict, bid: str) -> dict:
    return {
        "import_batch_id": bid,
        "device_name":  _to_str(row.get("Device_Name")),
        "device_ip":    _to_str(row.get("Device_IP")),
        "vendor":       _to_str(row.get("Vendor")),
        "router_id":    _to_str(row.get("Router_ID")),
        "ospf_admin":   _to_str(row.get("OSPF_Admin")),
        "if_ip":        _to_str(row.get("IF_IP")),
        "if_name":      _to_str(row.get("IF_Name")),
        "area":         _to_str(row.get("Area")),
        "if_admin":     _to_str(row.get("IF_Admin")),
        "if_state":     _to_str(row.get("IF_State")),
        "cost":         _to_int(row.get("Cost")),
        "mtu":          _to_int(row.get("MTU")),
        "data_source":  _to_str(row.get("Data_Source")),
        "captured_at":  _to_ts(row.get("Timestamp")),
    }


def _ospf_nbr_row(row: dict, bid: str) -> dict:
    return {
        "import_batch_id":      bid,
        "device_name":          _to_str(row.get("Device_Name")),
        "device_ip":            _to_str(row.get("Device_IP")),
        "vendor":               _to_str(row.get("Vendor")),
        "router_id":            _to_str(row.get("Router_ID")),
        "ospf_admin":           _to_str(row.get("OSPF_Admin")),
        "neighbor_ip":          _to_str(row.get("Neighbor_IP")),
        "neighbor_router_id":   _to_str(row.get("Neighbor_Router_ID")),
        "neighbor_device_name": _to_str(row.get("Neighbor_Device_Name")),
        "name_source":          _to_str(row.get("Name_Source")),
        "neighbor_state":       _to_str(row.get("Neighbor_State")),
        "captured_at":          _to_ts(row.get("Timestamp")),
    }


def _ospf_err_row(row: dict, bid: str) -> dict:
    return {
        "import_batch_id": bid,
        "device_name":  _to_str(row.get("Device_Name")),
        "device_ip":    _to_str(row.get("Device_IP")),
        "vendor":       _to_str(row.get("Vendor")),
        "error_type":   _to_str(row.get("Error_Type")),
        "error_detail": _to_str(row.get("Error_Detail")),
        "captured_at":  _to_ts(row.get("Timestamp")),
    }


def _bgp_sum_row(row: dict, bid: str) -> dict:
    return {
        "import_batch_id":  bid,
        "device_name":      _to_str(row.get("Device_Name")),
        "device_ip":        _to_str(row.get("Device_IP")),
        "vendor":           _to_str(row.get("Vendor")),
        "router_id":        _to_str(row.get("Router_ID")),
        "local_as":         _to_int(row.get("Local_AS")),
        "bgp_admin_state":  _to_str(row.get("BGP_Admin_State")),
        "bgp_oper_state":   _to_str(row.get("BGP_Oper_State")),
        "total_peers":      _to_int(row.get("Total_Peers")),
        "established":      _to_int(row.get("Established")),
        "not_established":  _to_int(row.get("Not_Established")),
        "vpnv4_rcvd":       _to_int(row.get("VPNv4_Rcvd")),
        "vpnv4_active":     _to_int(row.get("VPNv4_Active")),
        "status":           _to_str(row.get("Status")),
        "captured_at":      _to_ts(row.get("Timestamp")),
    }


def _bgp_nbr_row(row: dict, bid: str) -> dict:
    return {
        "import_batch_id":      bid,
        "device_name":          _to_str(row.get("Device_Name")),
        "device_ip":            _to_str(row.get("Device_IP")),
        "vendor":               _to_str(row.get("Vendor")),
        "router_id":            _to_str(row.get("Router_ID")),
        "local_as":             _to_int(row.get("Local_AS")),
        "neighbor_ip":          _to_str(row.get("Neighbor_IP")),
        "neighbor_device_name": _to_str(row.get("Neighbor_Device_Name")),
        "name_source":          _to_str(row.get("Name_Source")),
        "remote_as":            _to_int(row.get("Remote_AS")),
        "description":          _to_str(row.get("Description")),
        "bgp_group":            _to_str(row.get("BGP_Group")),
        "bgp_state":            _to_str(row.get("BGP_State")),
        "up_down":              _to_str(row.get("Up_Down")),
        "flaps":                _to_int(row.get("Flaps")),
        "last_error":           _to_str(row.get("Last_Error")),
        "hold_time":            _to_int(row.get("Hold_Time")),
        "vpnv4_rcvd":           _to_int(row.get("VPNv4_Rcvd")),
        "vpnv4_active":         _to_int(row.get("VPNv4_Active")),
        "anomaly":              _to_str(row.get("Anomaly")),
        "captured_at":          _to_ts(row.get("Timestamp")),
    }


def _bgp_err_row(row: dict, bid: str) -> dict:
    return {
        "import_batch_id": bid,
        "device_name": _to_str(row.get("Device_Name")),
        "device_ip":   _to_str(row.get("Device_IP")),
        "vendor":      _to_str(row.get("Vendor")),
        "error":       _to_str(row.get("Error")),
    }


def _hw_alarm_sum_row(row: dict, bid: str) -> dict:
    return {
        "import_batch_id": bid,
        "device_name":     _to_str(row.get("Device_Name")),
        "device_ip":       _to_str(row.get("Device_IP")),
        "vendor":          _to_str(row.get("Vendor")),
        "critical":        _to_int(row.get("Critical")),
        "major":           _to_int(row.get("Major")),
        "minor":           _to_int(row.get("Minor")),
        "power_status":    _to_str(row.get("Power_Status")),
        "fan_status":      _to_str(row.get("Fan_Status")),
        "max_temp":        _to_str(row.get("Max_Temp")),
        "temp_threshold":  _to_str(row.get("Temp_Threshold")),
        "overall_status":  _to_str(row.get("Overall_Status")),
        "captured_at":     _to_ts(row.get("Timestamp")),
    }


def _hw_alarm_det_row(row: dict, bid: str) -> dict:
    return {
        "import_batch_id": bid,
        "device_name":  _to_str(row.get("Device_Name")),
        "device_ip":    _to_str(row.get("Device_IP")),
        "vendor":       _to_str(row.get("Vendor")),
        "category":     _to_str(row.get("Category")),
        "severity":     _to_str(row.get("Severity")),
        "component":    _to_str(row.get("Component")),
        "status":       _to_str(row.get("Status")),
        "detail":       _to_str(row.get("Detail")),
        "captured_at":  _to_ts(row.get("Timestamp")),
    }


def _hw_alarm_err_row(row: dict, bid: str) -> dict:
    return {
        "import_batch_id": bid,
        "device_name": _to_str(row.get("Device_Name")),
        "device_ip":   _to_str(row.get("Device_IP")),
        "vendor":      _to_str(row.get("Vendor")),
        "error":       _to_str(row.get("Error")),
    }


# ==============================================================================
# Xây dựng bảng devices (merge từ inventory + OSPF interfaces + BGP summary)
# ==============================================================================

def _build_devices(bid: str, inv: list, ospf_intfs: list, bgp_sums: list, hw_alarm_sums: "list | None" = None) -> list:
    device_map: dict = {}

    def key(name, ip):
        return f"{name or ''}|{ip or ''}"

    for r in inv:
        if not r.get("device_name") and not r.get("device_ip"):
            continue
        k = key(r.get("device_name"), r.get("device_ip"))
        e = device_map.setdefault(k, {
            "device_name": r.get("device_name"),
            "device_ip":   r.get("device_ip"),
            "vendor": None, "loopback_ip": None,
            "router_id": None, "local_as": None, "sources": set(),
        })
        e["vendor"]      = e["vendor"]      or r.get("vendor")
        e["loopback_ip"] = e["loopback_ip"] or r.get("loopback_ip")
        e["sources"].add("inventory")

    for r in ospf_intfs:
        if not r.get("device_name") and not r.get("device_ip"):
            continue
        k = key(r.get("device_name"), r.get("device_ip"))
        e = device_map.setdefault(k, {
            "device_name": r.get("device_name"),
            "device_ip":   r.get("device_ip"),
            "vendor": None, "loopback_ip": None,
            "router_id": None, "local_as": None, "sources": set(),
        })
        e["vendor"]    = e["vendor"]    or r.get("vendor")
        e["router_id"] = e["router_id"] or r.get("router_id")
        e["sources"].add("ospf")

    for r in bgp_sums:
        if not r.get("device_name") and not r.get("device_ip"):
            continue
        k = key(r.get("device_name"), r.get("device_ip"))
        e = device_map.setdefault(k, {
            "device_name": r.get("device_name"),
            "device_ip":   r.get("device_ip"),
            "vendor": None, "loopback_ip": None,
            "router_id": None, "local_as": None, "sources": set(),
        })
        e["vendor"]    = e["vendor"]    or r.get("vendor")
        e["router_id"] = e["router_id"] or r.get("router_id")
        e["local_as"]  = e["local_as"]  or r.get("local_as")
        e["sources"].add("bgp")

    for r in (hw_alarm_sums or []):
        if not r.get("device_name") and not r.get("device_ip"):
            continue
        k = key(r.get("device_name"), r.get("device_ip"))
        e = device_map.setdefault(k, {
            "device_name": r.get("device_name"),
            "device_ip":   r.get("device_ip"),
            "vendor": None, "loopback_ip": None,
            "router_id": None, "local_as": None, "sources": set(),
        })
        e["vendor"] = e["vendor"] or r.get("vendor")
        e["sources"].add("hw_alarm")

    return [
        {
            "import_batch_id": bid,
            "device_name": e["device_name"],
            "device_ip":   e["device_ip"],
            "vendor":      e["vendor"],
            "loopback_ip": e["loopback_ip"],
            "router_id":   e["router_id"],
            "local_as":    e["local_as"],
            "source":      ",".join(sorted(e["sources"])),
        }
        for e in device_map.values()
    ]


# ==============================================================================
# IP helpers
# ==============================================================================

def _compute_network(ip: "str | None", prefix_len: "int | None") -> "str | None":
    """Python equivalent of Postgres: cidr(set_masklen(ip_address, prefix_length))."""
    if not ip or prefix_len is None:
        return None
    try:
        return str(ipaddress.IPv4Network(f"{ip}/{prefix_len}", strict=False))
    except ValueError:
        return None


def _ip_in_cidr(ip: "str | None", cidr: "str | None") -> bool:
    if not ip or not cidr:
        return False
    try:
        return ipaddress.ip_address(ip) in ipaddress.ip_network(cidr, strict=False)
    except ValueError:
        return False


def _enrich_network(rows: list) -> list:
    """Thêm trường 'network' (tính trong Python) vào mỗi dòng inventory."""
    enriched = []
    for r in rows:
        net = _compute_network(r.get("ip_address"), r.get("prefix_length"))
        enriched.append({**r, "network": net})
    return enriched


# ==============================================================================
# Audit Rule Engine
# ==============================================================================

_SEV_BASE = {"Critical": 100, "High": 75, "Medium": 45, "Low": 20, "Info": 10}
_LOWER_SEV = {
    "Critical": "High", "High": "Medium",
    "Medium": "Low", "Low": "Info", "Info": "Info",
}


def _prio(sev: str, active=False, routing=False, multi=False, dup=False) -> int:
    s = _SEV_BASE.get(sev, 0)
    if active:   s += 20
    if routing:  s += 15
    if multi:    s += 10
    if dup:      s += 20
    return s


def _q(value, default="?"):
    """Tương đương toán tử `value ?? default` của engine TS: chỉ thay thế khi
    value là None (giữ nguyên 0 và chuỗi rỗng). Dùng cho các chuỗi title/detail
    để khớp chính xác với scripts/lib/audit-rules.ts — `dict.get(key, default)`
    KHÔNG tương đương vì key luôn tồn tại nên None lọt qua thành chuỗi "None"."""
    return default if value is None else value


def _build_audit_findings(
    bid: str,
    ip_assignments: list,   # enriched with "network"
    bgp_summaries: list,
    bgp_neighbors: list,
    ospf_neighbors: list,
    ospf_errors: list,
    ospf_interfaces: list,
    prev_ospf_neighbors: list | None = None,
) -> list:
    findings = []

    ospf_ips = {r["if_ip"] for r in ospf_interfaces if r.get("if_ip")}
    bgp_ips  = {r["neighbor_ip"] for r in bgp_neighbors if r.get("neighbor_ip")}

    def routing_ctx(ip):
        return bool(ip) and (ip in ospf_ips or ip in bgp_ips)

    # ------------------------------------------------------------------
    # IP_DUP_ACTIVE_ACTIVE / IP_DUP_ACTIVE_MIXED / IP_DUP_INACTIVE
    # ------------------------------------------------------------------
    by_ip: dict = {}
    for r in ip_assignments:
        if r.get("ip_address"):
            by_ip.setdefault(r["ip_address"], []).append(r)

    for ip, rows in by_ip.items():
        if len(rows) < 2:
            continue
        active_rows = [r for r in rows if r.get("status") == "Active"]
        devices     = {r.get("device_name") for r in rows}
        network     = next((r.get("network") for r in rows if r.get("network")), None)
        rep         = active_rows[0] if active_rows else rows[0]

        def _svc_bucket(st):
            if st in ("Uplink", "Management"):
                return "Uplink/Management"
            return st or "(no service_type)"

        svc_counts: dict = {}
        for r in rows:
            k = _svc_bucket(r.get("service_type"))
            svc_counts[k] = svc_counts.get(k, 0) + 1
        same_vrf = any(c >= 2 for c in svc_counts.values())
        vrf_scope = "same-VRF" if same_vrf else "cross-VRF"

        if len(active_rows) >= 2:
            code = "IP_DUP_ACTIVE_ACTIVE"; sev = "Critical"; conf = 100
            title = f"Duplicate IP {ip} is Active on {len(active_rows)} interfaces"
        elif len(active_rows) == 1:
            code = "IP_DUP_ACTIVE_MIXED"; sev = "High"; conf = 90
            title = f"Duplicate IP {ip}: 1 Active + {len(rows)-1} other interface(s)"
        else:
            code = "IP_DUP_INACTIVE"; sev = "Medium"; conf = 75
            title = f"Duplicate IP {ip} appears on {len(rows)} inactive interfaces"

        severity = sev if same_vrf else _LOWER_SEV[sev]
        findings.append({
            "import_batch_id": bid, "severity": severity,
            "category": "IP Duplicate", "rule_code": code,
            "title":  f"{title} [{vrf_scope}]",
            "detail": "; ".join(
                f"{_q(r.get('device_name'))} / {_q(r.get('interface_name'))} "
                f"[VRF {_q(r.get('vrf_instance'), '—')}] ({_q(r.get('status'))})"
                for r in rows
            ),
            "device_name": rows[0].get("device_name"),
            "device_ip":   rows[0].get("device_ip"),
            "ip_address": ip, "network": network,
            "interface_name": None,
            "service_type":  rep.get("service_type"),
            "vrf_instance":  rep.get("vrf_instance"),
            "intf_status":   rep.get("status"),
            "confidence": conf,
            "priority_score": _prio(
                severity,
                active=bool(active_rows), routing=routing_ctx(ip),
                multi=len(devices) > 1, dup=same_vrf,
            ),
            "evidence": {
                "ip_address": ip, "vrf_scope": vrf_scope,
                "same_vrf_collision": same_vrf,
                "distinct_vrfs": list({r.get("vrf_instance") or "—" for r in rows}),
                "rows": [{
                    "device_name": r.get("device_name"), "device_ip": r.get("device_ip"),
                    "interface_name": r.get("interface_name"), "vrf_instance": r.get("vrf_instance"),
                    "vlan_id": r.get("vlan_id"), "status": r.get("status"),
                    "service_type": r.get("service_type"),
                } for r in rows],
            },
        })

    # ------------------------------------------------------------------
    # NETWORK_OVERUSED
    # ------------------------------------------------------------------
    by_net: dict = {}
    for r in ip_assignments:
        if r.get("network") and r.get("prefix_length") is not None:
            by_net.setdefault(r["network"], []).append(r)

    for network, rows in by_net.items():
        pl = rows[0].get("prefix_length")
        if pl in (30, 31):
            expected = 2
        elif pl == 32:
            expected = 1
        else:
            continue
        if len(rows) <= expected:
            continue
        active_rows = [r for r in rows if r.get("status") == "Active"]
        severity = "High" if active_rows else "Medium"
        devices  = {r.get("device_name") for r in rows}
        rep      = active_rows[0] if active_rows else rows[0]
        findings.append({
            "import_batch_id": bid, "severity": severity,
            "category": "Network", "rule_code": "NETWORK_OVERUSED",
            "title":  f"Network {network} has {len(rows)} endpoints (expected <= {expected})",
            "detail": "; ".join(
                f"{_q(r.get('device_name'))} / {_q(r.get('interface_name'))} "
                f"= {_q(r.get('ip_address'))} ({_q(r.get('status'))})"
                for r in rows
            ),
            "device_name": rows[0].get("device_name"), "device_ip": rows[0].get("device_ip"),
            "ip_address": None, "network": network, "interface_name": None,
            "service_type": rep.get("service_type"), "vrf_instance": rep.get("vrf_instance"),
            "intf_status": rep.get("status"), "confidence": 85,
            "priority_score": _prio(
                severity, active=bool(active_rows),
                routing=any(routing_ctx(r.get("ip_address")) for r in rows),
                multi=len(devices) > 1, dup=True,
            ),
            "evidence": {
                "network": network, "prefix_length": pl,
                "expected_endpoints": expected,
                "rows": [{"device_name": r.get("device_name"),
                          "interface_name": r.get("interface_name"),
                          "ip_address": r.get("ip_address"), "status": r.get("status")}
                         for r in rows],
            },
        })

    # ------------------------------------------------------------------
    # GATEWAY_OUTSIDE_SUBNET
    # ------------------------------------------------------------------
    for r in ip_assignments:
        gw = r.get("gateway"); net = r.get("network")
        if not gw or not net:
            continue
        if _ip_in_cidr(gw, net):
            continue
        findings.append({
            "import_batch_id": bid, "severity": "High",
            "category": "Gateway", "rule_code": "GATEWAY_OUTSIDE_SUBNET",
            "title":  f"Gateway {gw} is outside subnet {net}",
            "detail": (f"{_q(r.get('device_name'))} / {_q(r.get('interface_name'))}: "
                       f"IP {r.get('ip_address')}/{r.get('prefix_length')}, gateway {gw}"),
            "device_name": r.get("device_name"), "device_ip": r.get("device_ip"),
            "ip_address": r.get("ip_address"), "network": net,
            "interface_name": r.get("interface_name"),
            "service_type": r.get("service_type"), "vrf_instance": r.get("vrf_instance"),
            "intf_status": r.get("status"), "confidence": 95,
            "priority_score": _prio("High",
                active=r.get("status") == "Active", routing=routing_ctx(r.get("ip_address"))),
            "evidence": {
                "ip_address": r.get("ip_address"), "prefix_length": r.get("prefix_length"),
                "network": net, "gateway": gw,
            },
        })

    # ------------------------------------------------------------------
    # STATUS_STATE_MISMATCH
    # ------------------------------------------------------------------
    for r in ip_assignments:
        adm = (r.get("admin_state") or "").lower()
        opr = (r.get("oper_state") or "").lower()
        reason = None
        if r.get("status") == "Active" and opr == "down":
            reason = "Status is Active but operational state is Down"
        elif r.get("status") == "Admin-Down" and adm != "down":
            reason = f"Status is Admin-Down but admin state is {r.get('admin_state') or 'unknown'}"
        if not reason:
            continue
        findings.append({
            "import_batch_id": bid, "severity": "Medium",
            "category": "Status", "rule_code": "STATUS_STATE_MISMATCH",
            "title":  f"{_q(r.get('device_name'))} / {_q(r.get('interface_name'))}: {reason}",
            "detail": (f"admin_state={_q(r.get('admin_state'))}, "
                       f"oper_state={_q(r.get('oper_state'))}, status={_q(r.get('status'))}"),
            "device_name": r.get("device_name"), "device_ip": r.get("device_ip"),
            "ip_address": r.get("ip_address"), "network": r.get("network"),
            "interface_name": r.get("interface_name"),
            "service_type": r.get("service_type"), "vrf_instance": r.get("vrf_instance"),
            "intf_status": r.get("status"), "confidence": 80,
            "priority_score": _prio("Medium",
                active=r.get("status") == "Active", routing=routing_ctx(r.get("ip_address"))),
            "evidence": {
                "admin_state": r.get("admin_state"),
                "oper_state":  r.get("oper_state"),
                "status":      r.get("status"),
            },
        })

    # ------------------------------------------------------------------
    # PREFIX_SERVICE_MISMATCH
    # ------------------------------------------------------------------
    for r in ip_assignments:
        pl = r.get("prefix_length"); svc = r.get("service_type")
        if pl is None or not svc:
            continue
        reason = None
        if svc == "Loopback" and pl != 32:
            reason = f"Loopback interface expected /32, found /{pl}"
        elif svc == "Uplink" and pl not in (30, 31):
            reason = f"Uplink interface expected /30 or /31, found /{pl}"
        elif svc == "Management" and pl == 32:
            reason = "Management interface unexpectedly uses /32 (no room for a gateway)"
        if not reason:
            continue
        findings.append({
            "import_batch_id": bid, "severity": "Low",
            "category": "Prefix", "rule_code": "PREFIX_SERVICE_MISMATCH",
            "title":  f"{_q(r.get('device_name'))} / {_q(r.get('interface_name'))}: {reason}",
            "detail": reason,
            "device_name": r.get("device_name"), "device_ip": r.get("device_ip"),
            "ip_address": r.get("ip_address"), "network": r.get("network"),
            "interface_name": r.get("interface_name"),
            "service_type": svc, "vrf_instance": r.get("vrf_instance"),
            "intf_status": r.get("status"), "confidence": 60,
            "priority_score": _prio("Low",
                active=r.get("status") == "Active", routing=routing_ctx(r.get("ip_address"))),
            "evidence": {"service_type": svc, "prefix_length": pl},
        })

    # ------------------------------------------------------------------
    # BGP_PEER_NOT_ESTABLISHED
    # ------------------------------------------------------------------
    for n in bgp_neighbors:
        if n.get("bgp_state") == "Established":
            continue
        findings.append({
            "import_batch_id": bid, "severity": "High",
            "category": "BGP", "rule_code": "BGP_PEER_NOT_ESTABLISHED",
            "title":  f"{_q(n.get('device_name'))}: BGP peer {n.get('neighbor_ip')} is {_q(n.get('bgp_state'), 'unknown')}",
            "detail": (f"remote_as={_q(n.get('remote_as'))}, group={_q(n.get('bgp_group'))}, "
                       f"last_error={_q(n.get('last_error'))}"),
            "device_name": n.get("device_name"), "device_ip": n.get("device_ip"),
            "ip_address": n.get("neighbor_ip"), "network": None,
            "interface_name": None, "service_type": None,
            "vrf_instance": None, "intf_status": None, "confidence": 90,
            "priority_score": _prio("High", routing=True),
            "evidence": {
                "bgp_state": n.get("bgp_state"), "remote_as": n.get("remote_as"),
                "bgp_group": n.get("bgp_group"), "last_error": n.get("last_error"),
                "anomaly":   n.get("anomaly"),
            },
        })

    # ------------------------------------------------------------------
    # BGP_DEVICE_WARNING_ERROR
    # ------------------------------------------------------------------
    for s in bgp_summaries:
        if s.get("status") not in ("WARNING", "ERROR"):
            continue
        sev = "High" if s.get("status") == "ERROR" else "Medium"
        findings.append({
            "import_batch_id": bid, "severity": sev,
            "category": "BGP", "rule_code": "BGP_DEVICE_WARNING_ERROR",
            "title":  f"{_q(s.get('device_name'))}: BGP summary status is {s.get('status')}",
            "detail": (f"peers={_q(s.get('total_peers'))}, established={_q(s.get('established'))}, "
                       f"not_established={_q(s.get('not_established'))}"),
            "device_name": s.get("device_name"), "device_ip": s.get("device_ip"),
            "ip_address": s.get("router_id"), "network": None,
            "interface_name": None, "service_type": None,
            "vrf_instance": None, "intf_status": None, "confidence": 90,
            "priority_score": _prio(sev, routing=True),
            "evidence": {
                "status": s.get("status"), "total_peers": s.get("total_peers"),
                "established": s.get("established"), "not_established": s.get("not_established"),
                "vpnv4_rcvd": s.get("vpnv4_rcvd"), "vpnv4_active": s.get("vpnv4_active"),
            },
        })

    # ------------------------------------------------------------------
    # BGP_HIGH_FLAPS
    # ------------------------------------------------------------------
    for n in bgp_neighbors:
        flaps = n.get("flaps")
        if flaps is None:
            continue
        if flaps >= 100_000:
            sev = "High"
        elif flaps >= 1_000:
            sev = "Medium"
        else:
            continue
        findings.append({
            "import_batch_id": bid, "severity": sev,
            "category": "BGP", "rule_code": "BGP_HIGH_FLAPS",
            "title":  f"{_q(n.get('device_name'))}: BGP peer {n.get('neighbor_ip')} has {flaps} flaps",
            "detail": f"bgp_state={_q(n.get('bgp_state'))}, up_down={_q(n.get('up_down'))}",
            "device_name": n.get("device_name"), "device_ip": n.get("device_ip"),
            "ip_address": n.get("neighbor_ip"), "network": None,
            "interface_name": None, "service_type": None,
            "vrf_instance": None, "intf_status": None, "confidence": 85,
            "priority_score": _prio(sev, routing=True),
            "evidence": {"flaps": flaps, "bgp_state": n.get("bgp_state"), "up_down": n.get("up_down")},
        })

    # ------------------------------------------------------------------
    # BGP_LOW_ACTIVE_RATIO
    # ------------------------------------------------------------------
    for n in bgp_neighbors:
        rcvd = n.get("vpnv4_rcvd")
        active = n.get("vpnv4_active")
        if (rcvd or 0) > 0 and (active or 0) == 0:
            findings.append({
                "import_batch_id": bid, "severity": "Medium",
                "category": "BGP", "rule_code": "BGP_LOW_ACTIVE_RATIO",
                "title":  (f"{_q(n.get('device_name'))}: BGP peer {n.get('neighbor_ip')} "
                           f"received {rcvd} VPNv4 routes but 0 are active"),
                "detail": f"vpnv4_rcvd={rcvd}, vpnv4_active={active}",
                "device_name": n.get("device_name"), "device_ip": n.get("device_ip"),
                "ip_address": n.get("neighbor_ip"), "network": None,
                "interface_name": None, "service_type": None,
                "vrf_instance": None, "intf_status": None, "confidence": 75,
                "priority_score": _prio("Medium", routing=True),
                "evidence": {"vpnv4_rcvd": rcvd, "vpnv4_active": active},
            })

    # ------------------------------------------------------------------
    # OSPF_NEIGHBOR_NOT_FULL
    # ------------------------------------------------------------------
    for n in ospf_neighbors:
        if (n.get("neighbor_state") or "").lower() == "full":
            continue
        findings.append({
            "import_batch_id": bid, "severity": "High",
            "category": "OSPF", "rule_code": "OSPF_NEIGHBOR_NOT_FULL",
            "title":  (f"{_q(n.get('device_name'))}: OSPF neighbor {n.get('neighbor_ip')} "
                       f"is {_q(n.get('neighbor_state'), 'unknown')}"),
            "detail": (f"neighbor_router_id={_q(n.get('neighbor_router_id'))}, "
                       f"neighbor_device={_q(n.get('neighbor_device_name'))}"),
            "device_name": n.get("device_name"), "device_ip": n.get("device_ip"),
            "ip_address": n.get("neighbor_ip"), "network": None,
            "interface_name": None, "service_type": None,
            "vrf_instance": None, "intf_status": None, "confidence": 90,
            "priority_score": _prio("High", routing=True),
            "evidence": {
                "neighbor_state":       n.get("neighbor_state"),
                "neighbor_router_id":   n.get("neighbor_router_id"),
                "neighbor_device_name": n.get("neighbor_device_name"),
            },
        })

    # ------------------------------------------------------------------
    # OSPF_NEIGHBOR_DISAPPEARED
    # ------------------------------------------------------------------
    if prev_ospf_neighbors:
        curr_keys = {
            f"{(n.get('device_name') or '').lower()}|{(n.get('neighbor_ip') or '').lower()}"
            for n in ospf_neighbors
        }
        for p in prev_ospf_neighbors:
            key = f"{(p.get('device_name') or '').lower()}|{(p.get('neighbor_ip') or '').lower()}"
            if key in curr_keys:
                continue
            findings.append({
                "import_batch_id": bid, "severity": "High",
                "category": "OSPF", "rule_code": "OSPF_NEIGHBOR_DISAPPEARED",
                "title":  (f"{_q(p.get('device_name'))}: OSPF neighbor {p.get('neighbor_ip')}"
                           f"{' (' + p.get('neighbor_device_name') + ')' if p.get('neighbor_device_name') else ''}"
                           f" không còn trong batch mới"),
                "detail": (f"neighbor_router_id={_q(p.get('neighbor_router_id'))}, "
                           f"neighbor_device={_q(p.get('neighbor_device_name'))}, "
                           f"prev_state={_q(p.get('neighbor_state'))}"),
                "device_name": p.get("device_name"), "device_ip": p.get("device_ip"),
                "ip_address": p.get("neighbor_ip"), "network": None,
                "interface_name": None, "service_type": None,
                "vrf_instance": None, "intf_status": None, "confidence": 85,
                "priority_score": _prio("High", routing=True),
                "evidence": {
                    "prev_neighbor_state":       p.get("neighbor_state"),
                    "prev_neighbor_router_id":   p.get("neighbor_router_id"),
                    "prev_neighbor_device_name": p.get("neighbor_device_name"),
                },
            })

    # ------------------------------------------------------------------
    # OSPF_COLLECTION_ERROR
    # ------------------------------------------------------------------
    for e in ospf_errors:
        findings.append({
            "import_batch_id": bid, "severity": "Medium",
            "category": "OSPF", "rule_code": "OSPF_COLLECTION_ERROR",
            "title":  f"{_q(e.get('device_name'))}: OSPF collection error ({_q(e.get('error_type'), 'unknown')})",
            "detail": e.get("error_detail"),
            "device_name": e.get("device_name"), "device_ip": e.get("device_ip"),
            "ip_address": None, "network": None, "interface_name": None,
            "service_type": None, "vrf_instance": None, "intf_status": None, "confidence": 100,
            "priority_score": _prio("Medium", routing=True),
            "evidence": {"error_type": e.get("error_type"), "error_detail": e.get("error_detail")},
        })

    return findings


# ==============================================================================
# Resource Reclaim Engine
# ==============================================================================

_RECLAIM_STATUSES = {"Admin-Down", "Link-Down", "Up/No-Peer"}


def _build_resource_candidates(
    bid: str, ip_assignments: list, ospf_interfaces: list, bgp_neighbors: list
) -> list:
    ospf_ips = {r["if_ip"] for r in ospf_interfaces if r.get("if_ip")}
    bgp_ips  = {r["neighbor_ip"] for r in bgp_neighbors if r.get("neighbor_ip")}

    by_ip: dict = {}
    for r in ip_assignments:
        if r.get("ip_address"):
            by_ip.setdefault(r["ip_address"], []).append(r)

    candidates = []

    for r in ip_assignments:
        status   = r.get("status") or ""
        oper_dn  = (r.get("oper_state") or "").lower() == "down"
        if status not in _RECLAIM_STATUSES and not oper_dn:
            continue

        score = 0; bk: dict = {}

        if status == "Admin-Down": score += 30; bk["status_admin_down"] = 30
        if status == "Link-Down":  score += 25; bk["status_link_down"] = 25
        if status == "Up/No-Peer": score += 20; bk["status_up_no_peer"] = 20

        ip       = r.get("ip_address")
        has_ospf = bool(ip and ip in ospf_ips)
        has_bgp  = bool(ip and ip in bgp_ips)
        if not has_ospf: score += 20; bk["no_ospf_presence"] = 20
        if not has_bgp:  score += 20; bk["no_bgp_presence"]  = 20

        if not r.get("port_description"): score += 10; bk["no_port_description"] = 10

        if r.get("service_type") == "Loopback": score -= 30; bk["loopback_penalty"] = -30

        others      = [x for x in by_ip.get(ip or "", []) if x is not r]
        dup_active  = any(x.get("status") == "Active" for x in others)
        dup_inact   = bool(others) and not dup_active

        if dup_active: score -= 50; bk["dup_ip_active_elsewhere"] = -50
        elif dup_inact: score -= 15; bk["dup_ip_inactive_elsewhere"] = -15

        confidence = "High" if score >= 70 else ("Medium" if score >= 40 else "Low")
        if dup_active:
            confidence = "Low"

        reasons = []
        if status:  reasons.append(f"Trạng thái: {status}")
        if oper_dn and status != "Link-Down": reasons.append("Trạng thái vận hành (oper_state): Down")
        if not has_ospf: reasons.append("IP không xuất hiện trong cấu hình OSPF interface nào")
        if not has_bgp:  reasons.append("IP không xuất hiện trong cấu hình BGP neighbor nào")
        if not r.get("port_description"): reasons.append("Port description đang trống")
        if r.get("service_type") == "Loopback": reasons.append("Là Loopback (bị trừ điểm ưu tiên)")
        if dup_active:
            reasons.append("CẢNH BÁO: IP này đang trùng và Active ở interface/thiết bị khác — KHÔNG thu hồi IP này")
        elif dup_inact:
            reasons.append("IP này còn được khai báo ở (các) interface khác nhưng đều không Active")
        elif ip:
            reasons.append("IP chỉ được khai báo duy nhất tại đây (không trùng nơi khác) — an toàn để thu hồi")

        candidates.append({
            "import_batch_id": bid,
            "candidate_type":  status or "Oper-Down",
            "score":           score,
            "priority_score":  score,
            "confidence":      confidence,
            "reason":          "\n".join(reasons),
            "device_name":     r.get("device_name"),
            "device_ip":       r.get("device_ip"),
            "ip_address":      ip,
            "network":         r.get("network"),
            "interface_name":  r.get("interface_name"),
            "service_type":    r.get("service_type"),
            "current_status":  r.get("status"),
            "evidence": {
                "status": r.get("status"), "admin_state": r.get("admin_state"),
                "oper_state": r.get("oper_state"),
                "has_ospf_adjacency": has_ospf, "has_bgp_neighbor": has_bgp,
                "port_description": r.get("port_description"),
                "service_type": r.get("service_type"),
                "duplicate_ip": bool(others),
                "duplicate_ip_active_elsewhere": dup_active,
                "duplicate_ip_rows": [{
                    "device_name": x.get("device_name"), "interface_name": x.get("interface_name"),
                    "vrf_instance": x.get("vrf_instance"), "status": x.get("status"),
                } for x in others],
                "score_breakdown": bk,
            },
        })

    return candidates


# ==============================================================================
# Topology upload helper
# ==============================================================================

def _upload_and_archive_topology(topo_file: "str | None", archive_dir) -> int:
    """Upload topology HTML lên Supabase Storage rồi archive. Trả về số file đã archive."""
    if not topo_file:
        return 0
    log.info("Upload topology HTML lên Supabase Storage...")
    try:
        with open(topo_file, "rb") as f:
            content = f.read()
        _supabase.storage.from_("topology").upload(
            path="ospf_topology.html",
            file=content,
            file_options={"content-type": "text/html; charset=utf-8", "upsert": "true"},
        )
        log.info("  Upload OK (%d KB)", len(content) // 1024)
    except Exception as exc:
        log.warning("  Storage upload thất bại (không ảnh hưởng import): %s", exc)

    base = os.path.basename(topo_file)
    dest = archive_dir / base
    if dest.exists():
        stem, ext = os.path.splitext(base)
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S_%f")[:21]
        dest = archive_dir / f"{stem}_{stamp}{ext}"
    try:
        shutil.move(topo_file, str(dest))
        log.info("  → %s", dest.name)
        return 1
    except Exception as exc:
        log.warning("  Không thể di chuyển %s: %s", base, exc)
        return 0


# ==============================================================================
# Main
# ==============================================================================

def main() -> None:
    t0 = datetime.now()
    archive_dir = _watch_dir / ARCHIVE_SUBDIR
    archive_dir.mkdir(parents=True, exist_ok=True)

    log.info("=" * 60)
    log.info("Transmission Hub — Auto Data Import")
    log.info("Thư mục nguồn : %s", _watch_dir)
    log.info("Thư mục archive: %s", archive_dir)

    # ------------------------------------------------------------------ #
    # 1. Tìm file mới nhất theo từng prefix                               #
    # ------------------------------------------------------------------ #
    log.info("Tìm file mới nhất...")

    # Topology HTML tìm trước — tùy chọn, độc lập với Excel
    topo_files = sorted(glob.glob(str(_watch_dir / "ospf_topology_*.html")))
    topo_file: "str | None" = topo_files[-1] if topo_files else None
    if topo_file:
        log.info("  topology  : %s", os.path.basename(topo_file))
    else:
        log.info("  topology  : (không có file ospf_topology_*.html — bỏ qua)")

    try:
        inv_file  = _find_latest("ipvlan_inventory_")
        ospf_file = _find_latest("ospf_baseline_")
        bgp_file  = _find_latest("bgp_audit_")
        has_core = True
    except FileNotFoundError:
        inv_file = ospf_file = bgp_file = None
        has_core = False

    try:
        hw_alarm_file: "str | None" = _find_latest("hw_alarm_")
    except FileNotFoundError:
        hw_alarm_file = None

    if not has_core and not hw_alarm_file:
        log.error("Không tìm thấy file Excel nào để import (cần ít nhất core files hoặc hw_alarm).")
        _upload_and_archive_topology(topo_file, archive_dir)
        sys.exit(1)

    if has_core:
        log.info("  inventory : %s", os.path.basename(inv_file))
        log.info("  ospf      : %s", os.path.basename(ospf_file))
        log.info("  bgp       : %s", os.path.basename(bgp_file))
    else:
        log.info("  core files: (không có — chỉ import hw_alarm)")
    log.info("  hw_alarm  : %s", os.path.basename(hw_alarm_file) if hw_alarm_file else "(không có — bỏ qua)")

    # ------------------------------------------------------------------ #
    # 2. Đọc workbooks                                                    #
    # ------------------------------------------------------------------ #
    log.info("Đọc file Excel...")
    inv_rows: list = []
    ospf_intf_raw: list = []
    ospf_nbr_raw: list = []
    ospf_err_raw: list = []
    bgp_sum_raw: list = []
    bgp_nbr_raw: list = []
    bgp_err_raw: list = []

    if has_core:
        try:
            inv_rows      = _read_sheet(inv_file,  "IP VLAN Inventory")
            ospf_intf_raw = _read_sheet(ospf_file, "Interfaces")
            ospf_nbr_raw  = _read_sheet(ospf_file, "Neighbors")
            ospf_err_raw  = _read_sheet(ospf_file, "Errors")
            bgp_sum_raw   = _read_sheet(bgp_file,  "BGP_Summary")
            bgp_nbr_raw   = _read_sheet(bgp_file,  "BGP_Neighbors")
            bgp_err_raw   = _read_sheet(bgp_file,  "Errors")
        except Exception as exc:
            log.error("Lỗi đọc Excel: %s", exc)
            sys.exit(1)

    hwa_sum_raw: list = []
    hwa_det_raw: list = []
    hwa_err_raw: list = []
    if hw_alarm_file:
        try:
            hwa_sum_raw = _read_sheet(hw_alarm_file, "Alarm_Summary")
            hwa_det_raw = _read_sheet(hw_alarm_file, "Alarm_Details")
            hwa_err_raw = _read_sheet(hw_alarm_file, "Errors")
        except Exception as exc:
            log.warning("Lỗi đọc hw_alarm Excel (bỏ qua): %s", exc)
            hw_alarm_file = None

    if has_core:
        log.info("  inventory rows   : %d", len(inv_rows))
        log.info("  ospf interfaces  : %d", len(ospf_intf_raw))
        log.info("  ospf neighbors   : %d", len(ospf_nbr_raw))
        log.info("  ospf errors      : %d", len(ospf_err_raw))
        log.info("  bgp summaries    : %d", len(bgp_sum_raw))
        log.info("  bgp neighbors    : %d", len(bgp_nbr_raw))
        log.info("  bgp errors       : %d", len(bgp_err_raw))
    if hw_alarm_file:
        log.info("  hw_alarm summary : %d", len(hwa_sum_raw))
        log.info("  hw_alarm details : %d", len(hwa_det_raw))
        log.info("  hw_alarm errors  : %d", len(hwa_err_raw))

    # ------------------------------------------------------------------ #
    # HW-ALARM-ONLY MODE                                                  #
    # Khi chỉ có file hw_alarm (không có 3 core files), cập nhật hw_alarm #
    # vào batch completed gần nhất thay vì tạo batch mới.                 #
    # ------------------------------------------------------------------ #
    if not has_core and hw_alarm_file:
        log.info("Chế độ hw_alarm-only: cập nhật vào batch gần nhất...")
        try:
            resp = (
                _supabase.table("import_batches")
                .select("id")
                .eq("status", "completed")
                .order("completed_at", desc=True)
                .limit(1)
                .execute()
            )
            if not resp.data:
                log.error("Không tìm thấy batch completed nào để cập nhật hw_alarm.")
                sys.exit(1)
            target_bid = resp.data[0]["id"]
            log.info("  batch mục tiêu: %s", target_bid)

            hwa_sum_pl = [_hw_alarm_sum_row(r, target_bid) for r in hwa_sum_raw]
            hwa_det_pl = [_hw_alarm_det_row(r, target_bid) for r in hwa_det_raw]
            hwa_err_pl = [_hw_alarm_err_row(r, target_bid) for r in hwa_err_raw]

            log.info("Xóa hw_alarm cũ trong batch...")
            _supabase.table("hw_alarm_summary").delete().eq("import_batch_id", target_bid).execute()
            _supabase.table("hw_alarm_details").delete().eq("import_batch_id", target_bid).execute()
            _supabase.table("hw_alarm_errors").delete().eq("import_batch_id", target_bid).execute()

            log.info("Insert hw_alarm mới...")
            if hwa_sum_pl:
                _insert_chunked("hw_alarm_summary", hwa_sum_pl)
            if hwa_det_pl:
                _insert_chunked("hw_alarm_details", hwa_det_pl)
            if hwa_err_pl:
                _insert_chunked("hw_alarm_errors", hwa_err_pl)

            _supabase.table("import_batches").update({
                "hw_alarm_summary_rows": len(hwa_sum_pl),
                "hw_alarm_detail_rows":  len(hwa_det_pl),
                "hw_alarm_error_rows":   len(hwa_err_pl),
            }).eq("id", target_bid).execute()
            log.info("  hw_alarm: %d summary, %d details, %d errors",
                     len(hwa_sum_pl), len(hwa_det_pl), len(hwa_err_pl))

        except Exception as exc:
            log.error("Lỗi cập nhật hw_alarm: %s", exc)
            sys.exit(1)

        _upload_and_archive_topology(topo_file, archive_dir)
        log.info("Chuyển file hw_alarm vào thư mục archive...")
        moved = 0
        for fpath in _find_all("hw_alarm_"):
            base = os.path.basename(fpath)
            dest = archive_dir / base
            if dest.exists():
                stem, ext = os.path.splitext(base)
                stamp = datetime.now().strftime("%Y%m%d_%H%M%S_%f")[:21]
                dest = archive_dir / f"{stem}_{stamp}{ext}"
            try:
                shutil.move(fpath, str(dest))
                log.info("  → %s", dest.name)
                moved += 1
            except Exception as exc:
                log.warning("  Không thể di chuyển %s: %s", base, exc)

        elapsed = (datetime.now() - t0).total_seconds()
        log.info(
            "Hoàn tất hw_alarm-only trong %.1fs | batch %s | %d file đã archive.",
            elapsed, target_bid, moved,
        )
        return

    # ------------------------------------------------------------------ #
    # 3. Tạo import batch                                                 #
    #                                                                    #
    # Lưu ý: dữ liệu cũ KHÔNG bị xóa ở đây mà chỉ dọn SAU KHI import     #
    # thành công (bước 11). Nhờ vậy nếu import lỗi giữa chừng, dữ liệu    #
    # batch cũ vẫn còn nguyên thay vì bị mất trắng. Các view "latest_*"  #
    # luôn lấy batch mới nhất nên giao diện vẫn hiển thị đúng.            #
    # ------------------------------------------------------------------ #
    log.info("Tạo import batch...")
    source_files = [
        os.path.basename(inv_file),
        os.path.basename(ospf_file),
        os.path.basename(bgp_file),
    ]
    if hw_alarm_file:
        source_files.append(os.path.basename(hw_alarm_file))

    try:
        resp = _supabase.table("import_batches").insert({
            "source_label": BATCH_SOURCE_LABEL,
            "source_files": source_files,
            "status": "running",
        }).execute()
        batch_id: str = resp.data[0]["id"]
        log.info("  batch id: %s", batch_id)
    except Exception as exc:
        log.error("Lỗi tạo import batch: %s", exc)
        sys.exit(1)

    try:
        # ---------------------------------------------------------------- #
        # 5. Transform dữ liệu                                             #
        # ---------------------------------------------------------------- #
        inv_payload      = [_inv_row(r, batch_id)       for r in inv_rows]
        ospf_intf_pl     = [_ospf_intf_row(r, batch_id) for r in ospf_intf_raw]
        ospf_nbr_pl      = [_ospf_nbr_row(r, batch_id)  for r in ospf_nbr_raw]
        ospf_err_pl      = [_ospf_err_row(r, batch_id)  for r in ospf_err_raw]
        bgp_sum_pl       = [_bgp_sum_row(r, batch_id)   for r in bgp_sum_raw]
        bgp_nbr_pl       = [_bgp_nbr_row(r, batch_id)   for r in bgp_nbr_raw]
        bgp_err_pl       = [_bgp_err_row(r, batch_id)   for r in bgp_err_raw]

        hwa_sum_pl       = [_hw_alarm_sum_row(r, batch_id) for r in hwa_sum_raw]
        hwa_det_pl       = [_hw_alarm_det_row(r, batch_id) for r in hwa_det_raw]
        hwa_err_pl       = [_hw_alarm_err_row(r, batch_id) for r in hwa_err_raw]

        devices_pl       = _build_devices(batch_id, inv_payload, ospf_intf_pl, bgp_sum_pl, hwa_sum_pl or None)

        # Tính trước "network" trong Python để dùng cho rule engines
        # (network là GENERATED ALWAYS AS trong DB → không được include khi insert)
        inv_enriched     = _enrich_network(inv_payload)

        # ---------------------------------------------------------------- #
        # 6. Insert dữ liệu                                                #
        # ---------------------------------------------------------------- #
        log.info("Insert dữ liệu lên Supabase...")
        _insert_chunked("devices",         devices_pl)
        _insert_chunked("ospf_interfaces", ospf_intf_pl)
        _insert_chunked("ospf_neighbors",  ospf_nbr_pl)
        _insert_chunked("ospf_errors",     ospf_err_pl)
        _insert_chunked("bgp_summary",     bgp_sum_pl)
        _insert_chunked("bgp_neighbors",   bgp_nbr_pl)
        _insert_chunked("bgp_errors",      bgp_err_pl)
        _insert_chunked("ip_assignments",  inv_payload)   # không gửi network
        if hwa_sum_pl:
            _insert_chunked("hw_alarm_summary", hwa_sum_pl)
        if hwa_det_pl:
            _insert_chunked("hw_alarm_details", hwa_det_pl)
        if hwa_err_pl:
            _insert_chunked("hw_alarm_errors",  hwa_err_pl)

        # ---------------------------------------------------------------- #
        # 7. Audit rule engine                                              #
        # ---------------------------------------------------------------- #
        log.info("Lấy OSPF neighbor batch trước để so sánh cross-batch...")
        prev_ospf_nbrs: list = []
        try:
            prev_resp = (
                _supabase.table("import_batches")
                .select("id")
                .eq("status", "completed")
                .order("created_at", desc=True)
                .limit(1)
                .execute()
            )
            prev_bid = (prev_resp.data[0]["id"] if prev_resp.data else None)
            if prev_bid:
                page_size = 1000
                offset = 0
                while True:
                    nbr_resp = (
                        _supabase.table("ospf_neighbors")
                        .select("*")
                        .eq("import_batch_id", prev_bid)
                        .range(offset, offset + page_size - 1)
                        .execute()
                    )
                    batch = nbr_resp.data or []
                    prev_ospf_nbrs.extend(batch)
                    if len(batch) < page_size:
                        break
                    offset += page_size
                log.info("  batch trước %s: %d OSPF neighbors", prev_bid, len(prev_ospf_nbrs))
            else:
                log.info("  không có batch trước")
        except Exception as exc:
            log.warning("  Không thể lấy OSPF neighbor batch trước: %s", exc)

        log.info("Chạy audit rule engine...")
        findings = _build_audit_findings(
            bid=batch_id,
            ip_assignments=inv_enriched,   # có network tính sẵn
            bgp_summaries=bgp_sum_pl,
            bgp_neighbors=bgp_nbr_pl,
            ospf_neighbors=ospf_nbr_pl,
            ospf_errors=ospf_err_pl,
            ospf_interfaces=ospf_intf_pl,
            prev_ospf_neighbors=prev_ospf_nbrs,
        )
        log.info("  %d findings", len(findings))
        _insert_chunked("audit_findings", findings)

        # ---------------------------------------------------------------- #
        # 8. Resource reclaim engine                                        #
        # ---------------------------------------------------------------- #
        log.info("Chạy resource-reclaim engine...")
        candidates = _build_resource_candidates(
            bid=batch_id,
            ip_assignments=inv_enriched,
            ospf_interfaces=ospf_intf_pl,
            bgp_neighbors=bgp_nbr_pl,
        )
        log.info("  %d candidates", len(candidates))
        _insert_chunked("resource_candidates", candidates)

        # ---------------------------------------------------------------- #
        # 9. Finalize batch                                                 #
        # ---------------------------------------------------------------- #
        log.info("Finalize import batch...")
        _supabase.table("import_batches").update({
            "status":                  "completed",
            "inventory_rows":          len(inv_payload),
            "ospf_interface_rows":     len(ospf_intf_pl),
            "ospf_neighbor_rows":      len(ospf_nbr_pl),
            "ospf_error_rows":         len(ospf_err_pl),
            "bgp_summary_rows":        len(bgp_sum_pl),
            "bgp_neighbor_rows":       len(bgp_nbr_pl),
            "bgp_error_rows":          len(bgp_err_pl),
            "audit_finding_rows":      len(findings),
            "resource_candidate_rows": len(candidates),
            "hw_alarm_summary_rows":   len(hwa_sum_pl),
            "hw_alarm_detail_rows":    len(hwa_det_pl),
            "hw_alarm_error_rows":     len(hwa_err_pl),
            "completed_at":            datetime.now(timezone.utc).isoformat(),
        }).eq("id", batch_id).execute()

    except Exception as exc:
        msg = str(exc)
        log.error("Lỗi trong quá trình import: %s", msg)
        try:
            _supabase.table("import_batches").update(
                {"status": "failed", "notes": msg[:2000]}
            ).eq("id", batch_id).execute()
        except Exception:
            pass
        sys.exit(1)

    # ------------------------------------------------------------------ #
    # 11. Dọn dữ liệu cũ — chỉ chạy SAU KHI import mới đã hoàn tất.       #
    #     Giữ 2 batch mới nhất (batch vừa import + batch trước đó để         #
    #     Dashboard so sánh delta), xóa phần còn lại (bảng con tự cascade  #
    #     nhờ ON DELETE CASCADE). Lỗi ở đây không ảnh hưởng dữ liệu mới   #
    #     vừa import nên chỉ cảnh báo, không dừng chương trình.           #
    # ------------------------------------------------------------------ #
    log.info("Dọn dữ liệu batch cũ trên Supabase (giữ 2 batch mới nhất)...")
    try:
        all_batches = (
            _supabase.table("import_batches")
            .select("id")
            .order("completed_at", desc=True, nulls_last=True)
            .order("created_at", desc=True, nulls_last=True)
            .execute()
        )
        batch_ids = [row["id"] for row in (all_batches.data or [])]
        keep_ids = set(batch_ids[:2])
        delete_ids = [bid for bid in batch_ids if bid not in keep_ids]
        if delete_ids:
            for did in delete_ids:
                _supabase.table("import_batches").delete().eq("id", did).execute()
            log.info("  Đã xóa %d batch cũ, giữ lại %d batch.", len(delete_ids), len(keep_ids))
        else:
            log.info("  Không có batch cũ cần xóa.")
    except Exception as exc:
        log.warning("  Không thể xóa batch cũ (không ảnh hưởng dữ liệu mới): %s", exc)

    # ------------------------------------------------------------------ #
    # 12. Upload topology HTML + archive tất cả file                     #
    # ------------------------------------------------------------------ #
    moved = _upload_and_archive_topology(topo_file, archive_dir)

    log.info("Chuyển file Excel vào thư mục archive...")
    for prefix in ("ipvlan_inventory_", "ospf_baseline_", "bgp_audit_", "hw_alarm_"):
        for fpath in _find_all(prefix):
            base = os.path.basename(fpath)
            dest = archive_dir / base
            if dest.exists():
                stem, ext = os.path.splitext(base)
                stamp = datetime.now().strftime("%Y%m%d_%H%M%S_%f")[:21]
                dest = archive_dir / f"{stem}_{stamp}{ext}"
            try:
                shutil.move(fpath, str(dest))
                log.info("  → %s", dest.name)
                moved += 1
            except Exception as exc:
                log.warning("  Không thể di chuyển %s: %s", base, exc)

    elapsed = (datetime.now() - t0).total_seconds()
    log.info(
        "Hoàn tất trong %.1fs | batch %s | %d file đã archive.",
        elapsed, batch_id, moved,
    )


if __name__ == "__main__":
    main()
